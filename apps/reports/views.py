# apps/reports/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Count, Avg, Sum, Q
from datetime import datetime, timedelta
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# CORRECTED IMPORTS - Task from tasks app, not projects
from apps.projects.models import Project
from apps.tasks.models import Task
from apps.accounts.models import User

@login_required
def reports_dashboard(request):
    """Reports dashboard with analytics"""
    context = {
        'total_projects': Project.objects.count(),
        'completed_projects': Project.objects.filter(status='COMPLETED').count(),
        'total_tasks': Task.objects.count(),
        'completed_tasks': Task.objects.filter(status='COMPLETED').count(),
        'total_users': User.objects.count(),
        'active_users': User.objects.filter(is_active=True).count(),
    }
    return render(request, 'reports/reports_dashboard.html', context)

@login_required
def generate_project_report(request):
    """Generate project report in Excel format"""
    if request.method == 'POST':
        wb = Workbook()
        
        # Projects sheet
        ws_projects = wb.active
        ws_projects.title = "Projects"
        ws_projects.append(['Project Name', 'Status', 'Priority', 'Start Date', 'End Date', 
                           'Progress', 'Total Tasks', 'Completed Tasks', 'Budget'])
        
        projects = Project.objects.all()
        for project in projects:
            total_tasks = project.tasks.count()
            completed_tasks = project.tasks.filter(status='COMPLETED').count()
            
            ws_projects.append([
                project.name,
                project.get_status_display(),
                project.get_priority_display(),
                str(project.start_date),
                str(project.end_date),
                f"{project.progress}%",
                total_tasks,
                completed_tasks,
                f"${project.budget}"
            ])
        
        # Tasks sheet
        ws_tasks = wb.create_sheet("Tasks")
        ws_tasks.append(['Task Title', 'Project', 'Assigned To', 'Priority', 'Status', 
                        'Deadline', 'Estimated Hours', 'Actual Hours'])
        
        tasks = Task.objects.all()
        for task in tasks:
            ws_tasks.append([
                task.title,
                task.project.name,
                task.assigned_to.username,
                task.get_priority_display(),
                task.get_status_display(),
                str(task.deadline),
                float(task.estimated_hours),
                float(task.actual_hours)
            ])
        
        # Employee performance sheet
        ws_employees = wb.create_sheet("Employee Performance")
        ws_employees.append(['Employee', 'Department', 'Total Tasks', 'Completed Tasks', 
                            'Completion Rate', 'Avg Hours per Task'])
        
        for user in User.objects.filter(role='EMPLOYEE'):
            total_tasks = Task.objects.filter(assigned_to=user).count()
            completed_tasks = Task.objects.filter(assigned_to=user, status='COMPLETED').count()
            completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
            avg_hours = Task.objects.filter(assigned_to=user).aggregate(Avg('actual_hours'))['actual_hours__avg'] or 0
            
            ws_employees.append([
                user.get_full_name(),
                user.department,
                total_tasks,
                completed_tasks,
                f"{completion_rate:.1f}%",
                f"{avg_hours:.1f}"
            ])
        
        # Generate response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="project_report.xlsx"'
        wb.save(response)
        return response
    
    return redirect('reports_dashboard')

@login_required
def generate_task_report(request):
    """Generate task report in PDF format"""
    if request.method == 'POST':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="task_report.pdf"'
        
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "Task Report")
        
        # Date
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 80, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Task summary
        y = height - 120
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Task Summary")
        
        y -= 20
        p.setFont("Helvetica", 10)
        total_tasks = Task.objects.count()
        completed_tasks = Task.objects.filter(status='COMPLETED').count()
        pending_tasks = Task.objects.filter(status='PENDING').count()
        in_progress_tasks = Task.objects.filter(status='IN_PROGRESS').count()
        
        p.drawString(50, y, f"Total Tasks: {total_tasks}")
        y -= 15
        p.drawString(50, y, f"Completed: {completed_tasks}")
        y -= 15
        p.drawString(50, y, f"Pending: {pending_tasks}")
        y -= 15
        p.drawString(50, y, f"In Progress: {in_progress_tasks}")
        
        # List top tasks
        y -= 30
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Recent Tasks")
        
        y -= 20
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Title")
        p.drawString(300, y, "Status")
        p.drawString(400, y, "Deadline")
        
        y -= 15
        p.setFont("Helvetica", 9)
        recent_tasks = Task.objects.order_by('-created_at')[:20]
        
        for task in recent_tasks:
            if y < 50:
                p.showPage()
                y = height - 50
                p.setFont("Helvetica", 9)
            
            p.drawString(50, y, task.title[:40] if len(task.title) > 40 else task.title)
            p.drawString(300, y, task.get_status_display())
            p.drawString(400, y, task.deadline.strftime('%Y-%m-%d'))
            y -= 15
        
        p.save()
        return response
    
    return redirect('reports_dashboard')

@login_required
def employee_performance_report(request):
    """Generate employee performance report"""
    if request.method == 'POST':
        from django.db import models
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Performance"
        
        # Headers
        headers = ['Employee Name', 'Department', 'Designation', 'Total Tasks', 
                  'Completed Tasks', 'Pending Tasks', 'Completion Rate', 
                  'Avg Hours/Task', 'On-Time Completion']
        ws.append(headers)
        
        employees = User.objects.filter(role='EMPLOYEE')
        
        for employee in employees:
            tasks = Task.objects.filter(assigned_to=employee)
            total = tasks.count()
            completed = tasks.filter(status='COMPLETED').count()
            pending = tasks.filter(status='PENDING').count()
            completion_rate = (completed / total * 100) if total > 0 else 0
            avg_hours = tasks.aggregate(Avg('actual_hours'))['actual_hours__avg'] or 0
            
            # On-time completion (tasks completed before deadline)
            on_time = tasks.filter(
                status='COMPLETED', 
                completed_at__lte=models.F('deadline')
            ).count()
            on_time_rate = (on_time / completed * 100) if completed > 0 else 0
            
            ws.append([
                employee.get_full_name(),
                employee.department or '',
                employee.designation or '',
                total,
                completed,
                pending,
                f"{completion_rate:.1f}%",
                f"{avg_hours:.1f}",
                f"{on_time_rate:.1f}%"
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employee_performance.xlsx"'
        wb.save(response)
        return response
    
    return redirect('reports_dashboard')